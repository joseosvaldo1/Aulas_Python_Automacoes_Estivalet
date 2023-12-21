"""
Planilhas Excel (Parte II)
* Convertendo Planilhas Excel para Json:

"""
import json
import openpyxl

print("Exemplo_1: ")
path = ".\planilhas\products.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['Sheet1']

# Usando o método 'iter_rows()':
print("Usando o metodo 'iter_rows()'")
for row in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3, values_only=True):
    print(row)

print(" ")
print(30*'-')
print(" ")
# Usando o método 'iter_cols()':
print("Usando o metodo 'iter_cols()'")
for col in sheet.iter_cols(min_row=2, max_row=4, min_col=1, max_col=3, values_only=True):
    print(col)    

print(" ")
print("Exemplo_2: ")

products = {}
for row in sheet.iter_rows(min_row=2, max_row=7, min_col=1, max_col=4, values_only=True):
    product_ID = row[0]
    product={
        "name": row[1],
        "price": row[2],
        "available": row[3],      
    }
    products[product_ID] = product

print(json.dumps(products, indent=4))