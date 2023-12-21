"""
Planilhas Excel (Parte I):

Lendo Planilhas do Excel:

- A Biblioteca Open Pai Excel do Python permite ler e escrever arquivos 
Excel de uma maneira simples e fácil.É importante observar que essa 
biblioteca funciona apenas com versões do Excel 2010 e superiores. Mas 
também podemos utilizar o LibreOffice para instalar. Então nós vamos 
rodar o Pip Install (comando no terminal: python -m install openpyxl), 
o Open Excel, e aqui é uma maneira bastante simples de usar.

- Um documento deplanilha Excel é chamado de workbook. Um único workbook 
é salvado em um arquivo com extensão .xlsx. Cada workbook contém várias 
planilhas (sheets, também chamadas de worksheets). A planilha que está 
sendo visualizada no momento (ou a última visualizada antes de fechar o 
Excel) se chama planilha ativa (active sheet).
- Cada planilha contém colunas (acessadas por meio de letras que começam
com A) e linhas (acessadas por meio de números que começam com 1). Uma
caixa em uma coluna e linha em particular é chamada de célula (cell). Cada
célula pode conter um valor numérico ou um texto. A grade de células com
dados forma uma planilha.

"""
import openpyxl
print("Exemplo_1: ")
wb = openpyxl.load_workbook('.\planilhas\example.xlsx')
print(f"type of 'wb' = {type(wb)}") 
# O método 'load_workbook()' retorna um objeto do tipo 'workbook'.
# Esse método recebe como parâmetro o caminho (relativo ou absoluto)
# do arquivo .xlsx com o qual a gente vai trabalhar.
wb_sheetnames = wb.sheetnames 
# O método 'sheetnames' retorna uma lista com os nomes das planilhas.
print(f"wb.sheetnames = {wb_sheetnames} ")
print(" ")
print("Exemplo_2: ")
# Para usar uma das planilhas individualmente, devemos armazená-la em uma
# variável a qual será atribuído nome do objeto workbook seguido do nome
# exato da planilha entre colchetes.

sheet = wb['Sheet1'] # => Objeto do tipo 'worksheet'.
print(f"sheet_type = {type(sheet)}")
print(f"sheet = {sheet}")
print(f"sheet_title = {sheet.title}")
active_sheet = wb.active
print(f"active_sheet = {active_sheet}")
print(" ")
print("Exemplo_3:")
# Determinando a quantidade de linhas e colunas:
# Para linhas (row): Usamos o método 'max_row()'.
# Para colunas (columns): Usamos o método 'max_columns()'.
print(f"sheet_max_row = {sheet.max_row}")
print(f"sheet_max_column = {sheet.max_column}")
print(" ")
print("Exemplo_4: ")
# Obtendo valores específicos de células: 
sheet_A1 = sheet['A1'] # Objeto do tipo 'Cell'.
sheet_A1_value = sheet['A1'].value
print(f"type_of_sheet_A1 = {sheet_A1}")
print(f"sheet_A1_value = {sheet_A1_value }")
print(40*'-')
sheet_B1 = sheet['B1'] # Objeto do tipo 'Cell'.
sheet_B1_value = sheet['B1'].value
print(f"type_of_sheet_B1 = {sheet_B1}")
print(f"sheet_B1_value = {sheet_B1_value}")
print(40*'-')

cell_B1 = sheet['B1']
cell_C1 = sheet['C1']

print('B1: Row %s, Column %s is %s ' % (cell_B1.row, cell_B1.column, cell_B1.value))
print('C1: Row %s, Column %s is %s ' % (cell_C1.row, cell_C1.column, cell_C1.value))
print('C1: Cell %s is %s ' % (cell_C1.coordinate, cell_C1.value))
print(40*'-')
print(f"Cell_B1 = {sheet.cell(row=1,column=2).value}")
print(" ")
print("Exemplo_4: ")
# Impressão de um intervalo de células:
for i in range(1,8):
    print("Cell_[" + str(i) + "] = " + sheet.cell(row=i, column=2).value)

print(" ")
print("Exemplo_5: ")
# Acessando uma região da planilha por meio de tuplas:
range = tuple(sheet['A1:C3'])
print(f"range = {range}")
for row_of_cell_objects in range:
    print(" ")
    print("======BEGIN OF ROW=====")
    print(" ")
    for cell_objects in row_of_cell_objects:
        
        print(cell_objects.coordinate, cell_objects.value)
        
    print(" ")
    print("=======END OF ROW======")
    print(" ")

print(" ")
print(" ")
print("Exemplo_6:")
# Usando o desempacotamento de variáveis:
cells = sheet['A1:C3']
print("Coluna_A,               Coluna_B,            Coluna_C")
for c1, c2, c3 in cells:
    
    print(f"{c1.value},     {c2.value},             {c3.value}")

print(" ")
print("Exemplo_7: ")
# Para ler uma coluna ou linha inteira, podemos usar os atributos 
# 'row' and 'column' do objeto sheet os quais precisam ser 
# convertidos inicialmente em listas antes de passarem para um 
# looping:

# As colunas e linhas do 'sheet' começam do 0.=> 
# sheet.columns[1] = coluna B
coluna_B = list(sheet.columns)[1]
print(f'Coluna_B = {coluna_B}')
print(" ")
print("Coluna B")
print(" ")
for cell_objects in coluna_B:
    print(cell_objects.value)

print(" ")
print("Exemplo_8:")
linha_1 = list(sheet.rows)[0] 
# Como a contagem de listas inicia-se por 0, temos: linha_1 = sheet.raw[0]
print(f'Linha_1 = {linha_1}')
print(" ")
print("Linha_1")
print(" ")
for cell_objects in linha_1:
    print(cell_objects.value)