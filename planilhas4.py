"""
Planilhas Excel (Parte Iv): 

*Escrevendo Dados Nas Células da Planilha:


"""
import openpyxl
print("Exemplo_1:")
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = "Hello, World!"
# Escrevendo por atribuição como se fosse um dicionário:
print(f"sheet[A1] = {sheet['A1'].value}")

# Escrevendo usando o método 'sheet.cell()':
sheet.cell(row=2, column=2).value = 128
print(f"B2 = {sheet.cell(row=2, column=2).value }")

# Salvando o arquivo usando o método 'wb.save()':
wb.save(r'C:\Users\Usuário\Desktop\Osvaldo\Work Spaces\ws-python_estivalet\planilhas\nova3.xlsx')
print(" ")
print("Exemplo_2:")
sheet = wb.active
# Usando um lopping 'for':
rows = {
    (88,46,57),
    (23,59,78),
    (56,21,98),
    (24,18,43),
    (34,15,67),
}

for row in rows:
    sheet.append(row) 

wb.save(r'.\planilhas\append.xlsx')
print(" ")
print("Exemplo_3:") 
sheet = wb.active
rows = {
    ('Bananas', 3.56),
    ('Maçãs', 8.90),
    ('Pão', 2.45),
    ('Iorgute', 9.90),
    ('Carne', 29.00),
    ('Feijão', 6.78),
    ('Arroz', 5.78),
}

for row in rows:
    sheet.append(row)

cell = sheet.cell(row=15, column=2)
cell.value = '=SUM(B8:B14)'

wb.save(r'.\planilhas\formula.xlsx')