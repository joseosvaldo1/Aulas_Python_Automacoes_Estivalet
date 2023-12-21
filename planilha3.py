"""
Planilhas Excel (Parte III)

* Criando Planilhas Excel:

"""
import openpyxl
print("Exemplo_1:")
#Criando o workbook:
wb = openpyxl.Workbook()
print(f"wb.sheet = {wb.sheetnames}")
sheet = wb.active
sheet.title = "Planilha n 1"
print(f"sheet.title = {sheet.title}")

#Salvando o conjunto de planilhas com o método 'wb.save()':
wb.save('nova.xlsx')

#Criando novas planilhas com o método 'create_sheet':
wb.create_sheet()
# wb.create_sheet(index=0, title='First Sheet')
# wb.create_sheet(index=2, title='Third Sheet')
print(f"wb.sheet_names = {wb.sheetnames}")

#Deletando as planilhas criadas com o método 'del':
# del wb['Third Sheet']
# del wb['Sheet']

wb.save('nova2.xlsx')